import openpyxl
import json
import sys


def process_xlsx(file_path="spreadsheet.xlsx"):
    wb = openpyxl.load_workbook(file_path)

    print(f"File: {file_path}")
    print(f"Số tab: {len(wb.sheetnames)}")
    print(f"Danh sách tab: {wb.sheetnames}")
    print("=" * 60)

    all_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Tab: {sheet_name} ---")
        print(f"  Số dòng: {ws.max_row}, Số cột: {ws.max_column}")

        # Đọc header từ dòng đầu tiên
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        print(f"  Headers: {headers}")

        # Đọc dữ liệu từ dòng thứ 2 trở đi
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value
                rows.append(row_dict)

        all_data[sheet_name] = rows
        print(f"  Số phòng: {len(rows)}")

        # In vài dòng đầu
        for item in rows[:3]:
            print(f"    {item}")
        if len(rows) > 3:
            print(f"    ... và {len(rows) - 3} phòng nữa")

    wb.close()
    return all_data


if __name__ == "__main__":
    file_path = sys.argv[1] if len(sys.argv) > 1 else "spreadsheet.xlsx"
    data = process_xlsx(file_path)

    # Lưu ra file JSON
    output_file = file_path.rsplit(".", 1)[0] + ".json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\nĐã lưu dữ liệu ra: {output_file}")
